#!/usr/bin/env python3
"""Erzeuge 14 PDF-Anschreiben aus der Excel-Tabelle und der HTML-Vorlage."""
import openpyxl
import subprocess
import re
import unicodedata
from pathlib import Path

ROOT = Path('/Users/decentnodes/code/smartwandler-homepage/local')
TEMPLATE = ROOT / 'anschreiben-vorlage.html'
EXCEL = ROOT / 'Ingenieurbüros.xlsx'
OUT = ROOT / 'briefe'
OUT.mkdir(exist_ok=True)

CHROME = '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome'
DATE_DE = '25.06.2026'  # Dresden, Datum

# Manuelle Pflege Anrede-Daten (Geschlecht + Titel-Kurzform):
# Schlüssel: Volltext aus Excel-Spalte "Ansprechpartner".
PERSONS = {
    'Dr.-Ing. Steffen Winkler': ('Herr', 'Dr.', 'Winkler'),
    'Dipl.-Ing. Steffen Tost': ('Herr', '', 'Tost'),
    'Dr.-Ing. Lutz Vogt': ('Herr', 'Dr.', 'Vogt'),
    'Dipl.-Ing. Jens-Peter Groß': ('Herr', '', 'Groß'),
    'Dr.-Ing. René Kipper': ('Herr', 'Dr.', 'Kipper'),
    'Dipl.-Ing. Steffen Müller': ('Herr', '', 'Müller'),
    'Heiko Schubert': ('Herr', '', 'Schubert'),
    'Sebastian Hoffmeister': ('Herr', '', 'Hoffmeister'),
    'Björn Böhme': ('Herr', '', 'Böhme'),
    'Dr.-Ing. habil. Thomas Luckner': ('Herr', 'Dr.', 'Luckner'),
    'Dr.-Ing. Uli Uhlig': ('Herr', 'Dr.', 'Uhlig'),
    'Dipl.-Ing. Sören Glöckner': ('Herr', '', 'Glöckner'),
    'Dipl.-Ing. Torsten Schulz': ('Herr', '', 'Schulz'),
    'Julius Okon': ('Herr', '', 'Okon'),
    'Dipl.-Ing. Mike Kleineberg': ('Herr', '', 'Kleineberg'),
    'Dipl.-Ing. (FH) Jacob Kornack': ('Herr', '', 'Kornack'),
    'Dipl.-Ing. Sandra Oschütz': ('Frau', '', 'Oschütz'),
    'Dipl.-Ing. Arno Bidmon': ('Herr', '', 'Bidmon'),
    'Heiko Augsburg': ('Herr', '', 'Augsburg'),
    'Henning Liebau': ('Herr', '', 'Liebau'),
    'Mario Kühnel': ('Herr', '', 'Kühnel'),
}

# Für Firmenname=None: Ersatz (z.B. Privat-Adresse Einzelperson)
FIRMA_FALLBACK = {
    22: 'Ingenieurbüro Kornack',
}


def slugify(text: str) -> str:
    text = unicodedata.normalize('NFKD', text)
    text = text.encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^\w\s-]', '', text).strip().lower()
    text = re.sub(r'[-\s]+', '-', text)
    return text or 'brief'


def split_addr(addr: str):
    """'Straße 1, 01099 Dresden' → ('Straße 1', '01099 Dresden')"""
    parts = [p.strip() for p in addr.split(',', 1)]
    if len(parts) == 2:
        return parts[0], parts[1]
    return parts[0], ''


def anrede(persons_text: str) -> str:
    names = [n.strip() for n in persons_text.split(',')]
    parts = []
    for n in names:
        if n not in PERSONS:
            raise KeyError(f'Person nicht in PERSONS gepflegt: {n!r}')
        anrede_, titel, nachname = PERSONS[n]
        # "Sehr geehrter Herr" / "Sehr geehrte Frau"
        if anrede_ == 'Herr':
            stamm = 'Sehr geehrter Herr'
        else:
            stamm = 'Sehr geehrte Frau'
        rest = (titel + ' ' if titel else '') + nachname
        parts.append(f'{stamm} {rest}')
    # erste Großschreibung, restliche kleingeschrieben
    out = parts[0]
    for p in parts[1:]:
        out += ', ' + p[0].lower() + p[1:]
    return out + ','


def build_empfaenger_html(firma, persons_text, street, city):
    lines = []
    if firma:
        lines.append(firma)
    # Personen einzeln pro Zeile
    for n in [x.strip() for x in persons_text.split(',')]:
        lines.append(n)
    lines.append(street)
    lines.append(city)
    return '<br>\n      '.join(lines)


def generate_html(template: str, firma, persons_text, street, city, anrede_text):
    empfaenger_html = build_empfaenger_html(firma, persons_text, street, city)
    # Bild- und QR-Pfade von relativ → absolut (file://...) anpassen,
    # damit sie auch aus dem briefe/-Unterordner heraus auflösen.
    template = template.replace(
        'src="images/logo-white.png"',
        f'src="file://{ROOT}/images/logo-white.png"',
    )
    # Empfaenger-Block ersetzen
    old_empf = re.search(
        r'<div class="empfaenger">.*?</div>',
        template, re.DOTALL,
    ).group(0)
    new_empf = f'<div class="empfaenger">\n      {empfaenger_html}\n    </div>'
    html = template.replace(old_empf, new_empf)
    # Datum
    html = re.sub(
        r'<div class="datum">.*?</div>',
        f'<div class="datum">Dresden, {DATE_DE}</div>',
        html, count=1, flags=re.DOTALL,
    )
    # Anrede
    html = re.sub(
        r'<p class="anrede">.*?</p>',
        f'<p class="anrede">{anrede_text}</p>',
        html, count=1, flags=re.DOTALL,
    )
    return html


def html_to_pdf(html_path: Path, pdf_path: Path):
    cmd = [
        CHROME,
        '--headless=new',
        '--disable-gpu',
        '--no-pdf-header-footer',
        '--no-margins',
        f'--print-to-pdf={pdf_path}',
        f'file://{html_path}',
    ]
    res = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    if res.returncode != 0:
        raise RuntimeError(f'Chrome PDF fail: {res.stderr}')


def main():
    template = TEMPLATE.read_text(encoding='utf-8')
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb['vereidigte_Ingnieure']
    count = 0
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=7).value
        if not status or 'Brief Vorbereitung' not in str(status):
            continue
        firma = ws.cell(row=row, column=1).value or FIRMA_FALLBACK.get(row, '')
        addr = ws.cell(row=row, column=2).value
        persons = ws.cell(row=row, column=4).value
        street, city = split_addr(addr)
        an = anrede(persons)
        html = generate_html(template, firma, persons, street, city, an)

        slug_base = slugify(firma) if firma else slugify(persons.split(',')[0])
        slug = f'{count+1:02d}_{slug_base}'
        html_path = OUT / f'{slug}.html'
        pdf_path = OUT / f'{slug}.pdf'
        html_path.write_text(html, encoding='utf-8')
        html_to_pdf(html_path, pdf_path)
        print(f'[{count+1:02d}] {firma or persons}  →  {pdf_path.name}')
        print(f'     Anrede: {an}')
        count += 1
    print(f'\nFertig: {count} Briefe erzeugt in {OUT}')


if __name__ == '__main__':
    main()
